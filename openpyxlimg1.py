#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2017-07-10 16:54:01
# @Author  : ditto (969956574@qq.com)
# @Link    : https://github.com/dittoyy
# @Version : $Id$
from openpyxl import Workbook,load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.compat import range
from openpyxl.styles import Font, Fill
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles import NamedStyle, Font, Border, Side
# #insert an image
# wb = Workbook()
# ws=wb.active
# ws['A1']='You should see three logos below'
# img = Image('flo.jpg')
# ws.add_image(img, 'A1')
# wb.save('img1.xlsx')

# #group col
# wb = Workbook()
# ws = wb.create_sheet('groupcol',1)
# ws.column_dimensions.group('A','D', hidden=True)
# wb.save('group1.xlsx')

# #merge
# wb=Workbook()
# ws=wb.active
# ws.merge_cells('B1:C1')
# # ws.unmerge_cells('B1:C1')
# wb.save('merge1.xlsx')
# ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=4)
# ws.unmerge_cells(start_row=2,start_column=1,end_row=2,end_column=4)


# #fomular
# wb=Workbook()
# ws=wb.active
# ws["A1"] = "=SUM(1, 1)"
# wb.save("formula1.xlsx")


# #Using number_formats
# import datetime
# wb=Workbook()
# ws=wb.active
# ws['A1'] = datetime.datetime(2010, 7, 21)
# print ws['A1'].number_format#yyyy-mm-dd h:mm:ss
# wb.guess_types = True
# ws['B1'] = '3.52%'
# # wb.guess_types = False
# print ws['B1'].value#0.0352
# print ws['B1'].number_format#0%
# wb.save("numfor1.xlsx")#4%四舍五入了

# #add workbook
# wb=Workbook()
# ws1=wb.active
# ws1.title='dido'
# for i in range(1,40):
#     ws1.append(range(600))

# col = ws1.column_dimensions['A']
# col.font = Font(bold=True)
# row = ws1.row_dimensions[1]
# row.font = Font(underline="single")



# ws2=wb.create_sheet('first')
# ws2['B2']='19222229'
# ws2['B2'].font = Font(size=12)

# ws3 = wb.create_sheet(title="Data")
# # for row in range(10, 20):
# #     for col in range(27, 54):
# #         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))

# print(ws2['AA10'].value)
# wb.save(filename = '3sheet.xlsx')

# wb = load_workbook(filename = '3sheet.xlsx')
# sheet_ranges = wb['first']
# print(sheet_ranges['B1'].value)

# #Edit Page Setup
wb=Workbook()
ws=wb.active
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
ws.page_setup.fitToHeight = 0
ws.page_setup.fitToWidth = 1
#Creating a Named Style
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)
bd = Side(style='thick', color="000000")
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
#use st yle
wb.add_named_style(highlight)
ws['A1'].style = highlight
ws['D5'].style = 'highlight'

wb.save(filename = 'stylename1.xlsx')