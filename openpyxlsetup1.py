#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2017-07-11 09:49:51
# @Author  : ditto (969956574@qq.com)
# @Link    : https://github.com/dittoyy
# @Version : $Id$

from openpyxl.workbook import Workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

wb = Workbook()
ws = wb.active

wsprops = ws.sheet_properties
wsprops.tabColor = "1072BA"
wsprops.filterMode = False
wsprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
wsprops.outlinePr.summaryBelow = False
wsprops.outlinePr.applyStyles = True
wsprops.pageSetUpPr.autoPageBreaks = True
# The basic syntax for creating a formatting rule is:
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='EE1111', end_color='EE1111'))
rule = Rule(type='cellIs', dxf=dxf, formula=["10"])

#convenience function for creating ColorScale rules
from openpyxl.formatting.rule import ColorScaleRule
rule = ColorScaleRule(start_type='percentile', start_value=10, start_color='FFAA0000',
                        mid_type='percentile', mid_value=50, mid_color='FF0000AA',
                        end_type='percentile', end_value=90, end_color='FF00AA00')

#convenience function for creating IconSet rules
from openpyxl.formatting.rule import IconSetRule
rule = IconSetRule('5Arrows', 'percent', [10, 20, 30, 40, 50], showValue=None, percent=None, reverse=None)

#convenience function for creating DataBar rules
from openpyxl.formatting.rule import DataBarRule
rule = DataBarRule(start_type='percentile', start_value=10, end_type='percentile', end_value='90',
                color="FF638EC6", showValue="None", minLength=None, maxLength=None)

