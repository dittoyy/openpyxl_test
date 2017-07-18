#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Date    : 2017-07-10 14:12:18
# @Author  : ditto (969956574@qq.com)
# @Link    : https://github.com/dittoyy
# @Version : $Id$


from openpyxl import Workbook

if __name__ == "__main__":
    print("写excel简单示例")

    # 创建一个excel工作区
    wb = Workbook()

    # 激活当前工作簿
    ws = wb.active

    # 往单元格A1写入数据, 其他单元格写入类似
    ws['A1'] = "开源优测"

    # 写下一行数据，列表元素对应每一个单元格
    ws.append([1, 2, 3])

    # 写入时间类型到excel, python会自动将类型转换成excel的日期时间类型
    import datetime
    ws['A2'] = datetime.datetime.now()

    # 保存为excel文件
    wb.save(u"简单excel写入示例.xlsx")